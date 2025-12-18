import asyncio
import base64
import csv
import io
import json
import os
import re
import shutil
import uuid
from datetime import datetime, timedelta
from typing import List, Optional, Set

import logging
import requests
from fastapi import Depends, File, Form, HTTPException, UploadFile
from fastapi import FastAPI
from fastapi.security import OAuth2PasswordBearer
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, Response
from fastapi.staticfiles import StaticFiles
from sqlalchemy import inspect, or_, text
from sqlalchemy.orm import Session
from PIL import Image, ImageColor, ImageDraw, ImageFont
from jose import JWTError, jwt
from passlib.context import CryptContext
from openpyxl import load_workbook

# Optional Cloudinary dependency
try:
    import cloudinary
    import cloudinary.uploader

    CLOUDINARY_AVAILABLE = True
except ImportError:
    CLOUDINARY_AVAILABLE = False

BASE_DIR = os.path.dirname(__file__)
UPLOADS_DIR = os.path.join(BASE_DIR, "uploads")
UPLOADS_TEMPLATE_DIR = os.path.join(UPLOADS_DIR, "templates")
UPLOADS_MASK_DIR = os.path.join(UPLOADS_DIR, "masks")


def load_env(path: str = ".env"):
    """Simple .env loader to set environment variables locally."""
    env_path = os.path.join(os.path.dirname(BASE_DIR), path)
    if not os.path.exists(env_path):
        return
    with open(env_path, "r") as f:
        for line in f:
            stripped = line.strip()
            if not stripped or stripped.startswith("#") or "=" not in stripped:
                continue
            key, value = stripped.split("=", 1)
            os.environ.setdefault(key.strip(), value.strip())


# Load environment early so Cloudinary config sees it
load_env()

from db import Base, SessionLocal, engine, get_session  # noqa: E402
from models import Job, ProcessedCompany, Template, User  # noqa: E402

logging.basicConfig(level=os.getenv("LOG_LEVEL", "INFO"))
logger = logging.getLogger(__name__)

CLOUD_NAME = os.getenv("CLOUDINARY_CLOUD_NAME", "")
CLOUD_API_KEY = os.getenv("CLOUDINARY_API_KEY", "")
CLOUD_API_SECRET = os.getenv("CLOUDINARY_API_SECRET", "")
PUBLIC_ID_PATTERN = os.getenv("PUBLIC_ID_PATTERN", "<template>-<row>-<uuid>")
UPLOAD_FOLDER = os.getenv("UPLOAD_FOLDER", "mockups")
SECRET_KEY = os.getenv("AUTH_SECRET_KEY") or os.getenv("SECRET_KEY")
if not SECRET_KEY:
    SECRET_KEY = base64.urlsafe_b64encode(os.urandom(32)).decode("utf-8")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = int(os.getenv("ACCESS_TOKEN_EXPIRE_MINUTES", "4320"))  # 3 days by default

if CLOUDINARY_AVAILABLE and CLOUD_NAME and CLOUD_API_KEY and CLOUD_API_SECRET:
    cloudinary.config(
        cloud_name=CLOUD_NAME,
        api_key=CLOUD_API_KEY,
        api_secret=CLOUD_API_SECRET,
        secure=True,
    )

# Ensure upload folder exists before mounting static files
os.makedirs(UPLOADS_TEMPLATE_DIR, exist_ok=True)

app = FastAPI()

# Allow frontend to connect
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
app.mount("/static", StaticFiles(directory=os.path.join(BASE_DIR, "template")), name="static")
app.mount("/uploads", StaticFiles(directory=UPLOADS_DIR), name="uploads")

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/auth/login")
pwd_context = CryptContext(
    schemes=["argon2"],
    deprecated="auto",
    argon2__memory_cost=15360,
    argon2__time_cost=2,
    argon2__parallelism=1,
)


# === Helpers ===
def shorten_company_name(name: str) -> str:
    suffixes = ["inc", "llc", "ltd", "co", "company", "group", "plc", "corp", "corporation"]
    words = name.strip().split()
    if words and words[-1].lower().strip(",.") in suffixes:
        words = words[:-1]
    cleaned = " ".join(words)
    cleaned = re.sub(r"[^a-zA-Z0-9 ]+", "", cleaned).strip()
    return " ".join(word.capitalize() for word in cleaned.split())


def ensure_upload_dir():
    if not os.path.exists(UPLOADS_TEMPLATE_DIR):
        os.makedirs(UPLOADS_TEMPLATE_DIR, exist_ok=True)
    if not os.path.exists(UPLOADS_MASK_DIR):
        os.makedirs(UPLOADS_MASK_DIR, exist_ok=True)


def run_migrations(engine):
    """Apply lightweight in-place migrations for existing databases."""
    insp = inspect(engine)
    with engine.begin() as conn:
        try:
            template_cols = {c["name"] for c in insp.get_columns("templates")}
        except Exception:
            template_cols = set()
        if "overlays" not in template_cols:
            if engine.dialect.name == "postgresql":
                conn.execute(text("ALTER TABLE templates ADD COLUMN overlays JSONB DEFAULT '[]'::jsonb"))
            else:
                conn.execute(text("ALTER TABLE templates ADD COLUMN overlays JSON DEFAULT '[]'"))
        if "masks" not in template_cols:
            if engine.dialect.name == "postgresql":
                conn.execute(text("ALTER TABLE templates ADD COLUMN masks JSONB DEFAULT '[]'::jsonb"))
            else:
                conn.execute(text("ALTER TABLE templates ADD COLUMN masks JSON DEFAULT '[]'"))
        if "owner_id" not in template_cols:
            conn.execute(text("ALTER TABLE templates ADD COLUMN owner_id TEXT"))

        try:
            job_cols = {c["name"] for c in insp.get_columns("jobs")}
        except Exception:
            job_cols = set()
        if "skip_processed" not in job_cols:
            if engine.dialect.name == "postgresql":
                conn.execute(text("ALTER TABLE jobs ADD COLUMN skip_processed BOOLEAN DEFAULT FALSE"))
            else:
                conn.execute(text("ALTER TABLE jobs ADD COLUMN skip_processed BOOLEAN DEFAULT 0"))
        if "identifier_column" not in job_cols:
            conn.execute(text("ALTER TABLE jobs ADD COLUMN identifier_column TEXT"))
        if "owner_id" not in job_cols:
            conn.execute(text("ALTER TABLE jobs ADD COLUMN owner_id TEXT"))


def hash_password(password: str) -> str:
    if not password or len(password) < 8:
        raise HTTPException(status_code=400, detail="Password must be at least 8 characters long.")
    return pwd_context.hash(password)


def verify_password(password: str, hashed: str) -> bool:
    if not password or not hashed:
        return False
    return pwd_context.verify(password, hashed)


def create_access_token(user_id: str, expires_minutes: int = ACCESS_TOKEN_EXPIRE_MINUTES) -> str:
    expire = datetime.utcnow() + timedelta(minutes=expires_minutes)
    payload = {"sub": user_id, "exp": expire, "iat": datetime.utcnow()}
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)


def decode_user_id(token: str) -> Optional[str]:
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        return payload.get("sub")
    except JWTError:
        return None


def user_to_dict(user: User) -> dict:
    return {
        "id": user.id,
        "email": user.email,
        "name": user.name,
        "created_at": user.created_at.isoformat() if user.created_at else None,
    }


def get_user_by_email(db: Session, email: str) -> Optional[User]:
    return db.query(User).filter(User.email == email.lower().strip()).first()


async def require_user(token: str = Depends(oauth2_scheme), db: Session = Depends(get_session)) -> User:
    user_id = decode_user_id(token)
    if not user_id:
        raise HTTPException(status_code=401, detail="Invalid or expired credentials.")
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=401, detail="Invalid or expired credentials.")
    return user

def resolve_template_image_path(path: str) -> str:
    """Return an absolute filesystem path for the template image."""
    # Already absolute and exists
    if os.path.isabs(path) and os.path.exists(path):
        return path
    # If path is served from /static/, map to template directory
    cleaned = path.lstrip("/")
    candidate = os.path.join(BASE_DIR, cleaned)
    if os.path.exists(candidate):
        return candidate
    candidate = os.path.join(BASE_DIR, "template", os.path.basename(path))
    if os.path.exists(candidate):
        return candidate
    return path


def default_template():
    """Seed a default template using the existing mockup asset."""
    return {
        "id": "default",
        "name": "Default Mockup",
        "baseImagePath": "/static/mockup-template.jpg",
        "overlays": [],
        "masks": [],
        "variables": [
            {
                "id": "short_name",
                "label": "Short Name",
                "type": "text",
                "x": 406,
                "y": 179,
                "w": 600,
                "h": 110,
                "style": {
                    "font": "Inter_Bold",
                    "size": 48,
                    "weight": "bold",
                    "color": "#000000",
                    "align": "left",
                },
                "defaultValue": "",
            },
            {
                "id": "full_name",
                "label": "Full Name",
                "type": "text",
                "x": 515,
                "y": 509,
                "w": 600,
                "h": 80,
                "style": {
                    "font": "Inter_Bold",
                    "size": 36,
                    "weight": "bold",
                    "color": "#bbbbbb",
                    "align": "left",
                },
                "defaultValue": "",
            },
            {
                "id": "logo",
                "label": "Logo",
                "type": "image",
                "x": 100,
                "y": 200,
                "w": 200,
                "h": 200,
                "fit": "contain",
                "defaultValue": "",
            },
        ],
    }


def roofing_template():
    """Seed a roofing template if the asset exists."""
    return {
        "id": "roofing",
        "name": "Roofing Mockup",
        "baseImagePath": "/static/roofing-template.jpg",
        "overlays": [],
        "masks": [],
        "variables": [
            {
                "id": "short_name",
                "label": "Short Name",
                "type": "text",
                "x": 406,
                "y": 179,
                "w": 600,
                "h": 110,
                "style": {
                    "font": "Inter_Bold",
                    "size": 48,
                    "weight": "bold",
                    "color": "#000000",
                    "align": "left",
                },
                "defaultValue": "",
            },
            {
                "id": "full_name",
                "label": "Full Name",
                "type": "text",
                "x": 515,
                "y": 509,
                "w": 600,
                "h": 80,
                "style": {
                    "font": "Inter_Bold",
                    "size": 36,
                    "weight": "bold",
                    "color": "#bbbbbb",
                    "align": "left",
                },
                "defaultValue": "",
            },
            {
                "id": "logo",
                "label": "Logo",
                "type": "image",
                "x": 100,
                "y": 200,
                "w": 200,
                "h": 200,
                "fit": "contain",
                "defaultValue": "",
            },
        ],
    }


def plumber_template():
    """Seed a plumber/handyman template if the asset exists."""
    return {
        "id": "plumber-1",
        "name": "Plumber Mockup",
        "baseImagePath": "/static/plumber-template-1.png",
        "overlays": [],
        "masks": [],
        "variables": [
            {
                "id": "short_name",
                "label": "Short Name",
                "type": "text",
                "x": 90,
                "y": 200,
                "w": 720,
                "h": 200,
                "style": {
                    "font": "Inter_Bold",
                    "size": 64,
                    "weight": "bold",
                    "color": "#ffffff",
                    "align": "left",
                },
                "defaultValue": "",
            },
            {
                "id": "full_name",
                "label": "Full Name",
                "type": "text",
                "x": 90,
                "y": 420,
                "w": 660,
                "h": 140,
                "style": {
                    "font": "Inter_Regular",
                    "size": 32,
                    "weight": "bold",
                    "color": "#d6e2ff",
                    "align": "left",
                },
                "defaultValue": "",
            },
            {
                "id": "logo",
                "label": "Logo",
                "type": "image",
                "x": 90,
                "y": 90,
                "w": 200,
                "h": 120,
                "fit": "contain",
                "defaultValue": "",
            },
        ],
    }


def get_font(style: dict):
    font_name = style.get("font", "Inter_Bold")
    size = style.get("size", 32)
    font_map = {
        "Inter_Bold": os.path.join(BASE_DIR, "fonts", "Inter_24pt-Bold.ttf"),
        "Inter_Regular": os.path.join(BASE_DIR, "fonts", "Inter_24pt-Regular.ttf"),
    }
    font_path = font_map.get(font_name, font_map["Inter_Bold"])
    try:
        return ImageFont.truetype(font_path, size)
    except Exception:
        return ImageFont.load_default()


def wrap_text_to_box(draw: ImageDraw.ImageDraw, text: str, font, max_width: int):
    words = text.split()
    lines = []
    current = ""
    for word in words:
        test_line = f"{current} {word}".strip()
        bbox = draw.textbbox((0, 0), test_line, font=font)
        width = bbox[2] - bbox[0]
        if width <= max_width:
            current = test_line
        else:
            if current:
                lines.append(current)
            current = word
    if current:
        lines.append(current)
    return lines


def draw_text_block(draw: ImageDraw.ImageDraw, variable: dict, text: str):
    x, y, w, h = variable["x"], variable["y"], variable["w"], variable["h"]
    style = variable.get("style", {})
    font = get_font(style)
    align = style.get("align", "left")
    valign = style.get("valign", "middle")

    lines = wrap_text_to_box(draw, text, font, w)
    # Measure each line to derive accurate height/width
    line_heights = []
    for line in lines:
        bbox = draw.textbbox((0, 0), line or "Ay", font=font)
        line_heights.append(bbox[3] - bbox[1])
    line_height = max(line_heights) if line_heights else font.getbbox("Ay")[3] - font.getbbox("Ay")[1]
    total_height = len(lines) * line_height

    valign_norm = valign.lower()
    if valign_norm == "top":
        start_y = y
    elif valign_norm == "bottom":
        start_y = y + max(h - total_height, 0)
    else:  # middle/default
        start_y = y + max((h - total_height) // 2, 0)
    # Upward nudge (vertical only) to better match on-screen editor preview
    start_y = max(start_y - int(line_height * 0.18), y)

    for i, line in enumerate(lines):
        bbox = draw.textbbox((0, 0), line, font=font)
        line_width = bbox[2] - bbox[0]
        if align == "center":
            line_x = x + max((w - line_width) // 2, 0)
        elif align == "right":
            line_x = x + max(w - line_width, 0)
        else:
            line_x = x
        draw.text((line_x, start_y + i * line_height), line, font=font, fill=style.get("color", "#000000"))


def load_image_from_value(value: str) -> Optional[Image.Image]:
    if not value:
        return None
    try:
        if value.startswith("data:image"):
            header, b64data = value.split(",", 1)
            binary = base64.b64decode(b64data)
            return Image.open(io.BytesIO(binary)).convert("RGB")
        if value.startswith("http://") or value.startswith("https://"):
            resp = requests.get(value, timeout=10)
            resp.raise_for_status()
            return Image.open(io.BytesIO(resp.content)).convert("RGB")
        if os.path.exists(value):
            return Image.open(value).convert("RGB")
    except Exception:
        return None
    return None


def load_mask_image(mask_entry) -> Optional[Image.Image]:
    """Return a mask image from either inline base64 data or a stored path."""
    data_value = None
    path_value = None
    if isinstance(mask_entry, dict):
        data_value = mask_entry.get("data")
        path_value = mask_entry.get("path")
    else:
        path_value = mask_entry

    if data_value:
        try:
            _, b64data = data_value.split(",", 1)
            binary = base64.b64decode(b64data)
            return Image.open(io.BytesIO(binary)).convert("RGBA")
        except Exception:
            return None

    if path_value:
        resolved = resolve_template_image_path(path_value)
        if os.path.exists(resolved):
            try:
                return Image.open(resolved).convert("RGBA")
            except Exception:
                return None
    return None


def paste_image(draw_image: Image.Image, variable: dict, value: str):
    img = load_image_from_value(value)
    if not img:
        return
    x, y, w, h = variable["x"], variable["y"], variable["w"], variable["h"]
    fit_mode = variable.get("fit", "cover")

    if fit_mode == "contain":
        img.thumbnail((w, h))
        paste_x = x + (w - img.width) // 2
        paste_y = y + (h - img.height) // 2
        draw_image.paste(img, (paste_x, paste_y))
    else:
        img_ratio = img.width / img.height
        box_ratio = w / h
        if img_ratio > box_ratio:
            new_width = w
            new_height = int(w / img_ratio)
        else:
            new_height = h
            new_width = int(h * img_ratio)
        img = img.resize((new_width, new_height))
        paste_x = x + (w - new_width) // 2
        paste_y = y + (h - new_height) // 2
        draw_image.paste(img, (paste_x, paste_y))


def template_to_dict(tpl: Template) -> dict:
    return {
        "id": tpl.id,
        "name": tpl.name,
        "baseImagePath": tpl.base_image_path,
        "variables": tpl.variables or [],
        "masks": tpl.masks or [],
        "ownerId": tpl.owner_id,
    }


def job_to_dict(job: Job) -> dict:
    return {
        "id": job.id,
        "status": job.status,
        "progress": job.progress,
        "created_at": job.created_at.isoformat() if job.created_at else None,
        "results": job.results or [],
        "rows": job.rows or [],
        "mapping": job.mapping or {},
        "template_id": job.template_id,
        "csv_path": job.csv_path,
        "skip_processed": bool(job.skip_processed),
        "identifier_column": job.identifier_column,
        "ownerId": job.owner_id,
    }


def apply_variables(template: dict, row: dict, mapping: dict) -> Image.Image:
    base_path = resolve_template_image_path(template["baseImagePath"])
    if not os.path.exists(base_path):
        raise FileNotFoundError(f"Template image not found: {base_path}")
    image = Image.open(base_path).convert("RGBA")

    masks = template.get("masks") or []
    if masks:
        mask_layer = Image.new("RGBA", image.size, (0, 0, 0, 0))
        for mask in masks:
            mask_img = load_mask_image(mask)
            if not mask_img:
                continue
            mask_img = mask_img.resize(image.size)
            mask_layer = Image.alpha_composite(mask_layer, mask_img)
        image = Image.alpha_composite(image, mask_layer)

    draw = ImageDraw.Draw(image)

    for variable in template.get("variables", []):
        var_id = variable["id"]
        column = mapping.get(var_id)
        value = ""
        if column:
            value = str(row.get(column, "")).strip()
        if not value:
            value = variable.get("defaultValue", "")
        if variable["type"] == "text":
            if not value and var_id == "short_name":
                value = shorten_company_name(str(row.get(column, "")))
            draw_text_block(draw, variable, value)
        elif variable["type"] == "image":
            paste_image(image, variable, value)
    return image.convert("RGB")


def public_id_for_row(template_id: str, row_index: int, row: dict) -> str:
    company = row.get("Company Name") or row.get("company") or "mockup"
    safe_company = re.sub(r"[^a-zA-Z0-9_-]+", "_", str(company)).strip("_") or "mockup"
    uid = uuid.uuid4().hex[:8]
    replacements = {
        "<template>": template_id,
        "<row>": str(row_index),
        "<company>": safe_company,
        "<uuid>": uid,
        "<ts>": datetime.utcnow().strftime("%Y%m%d%H%M%S"),
    }
    pid = PUBLIC_ID_PATTERN
    for key, val in replacements.items():
        pid = pid.replace(key, val)
    return pid


def upload_to_cloudinary(image: Image.Image, template_id: str, row_index: int, row: dict) -> str:
    if not CLOUDINARY_AVAILABLE:
        raise RuntimeError("Cloudinary SDK not installed")
    if not CLOUD_NAME or not CLOUD_API_KEY or not CLOUD_API_SECRET:
        raise RuntimeError("Cloudinary credentials missing")
    buffer = io.BytesIO()
    image.save(buffer, format="JPEG", quality=80, optimize=True)
    buffer.seek(0)
    public_id = public_id_for_row(template_id, row_index, row)
    result = cloudinary.uploader.upload(
        file=buffer.getvalue(),
        folder=UPLOAD_FOLDER,
        public_id=public_id,
        overwrite=True,
        resource_type="image",
        use_filename=False,
    )
    return result.get("secure_url") or result.get("url")


def parse_csv_file(file: UploadFile) -> List[dict]:
    content = file.file.read()
    file.file.seek(0)
    filename = (file.filename or "").lower()
    content_type = (file.content_type or "").lower()
    is_excel = filename.endswith((".xlsx", ".xls")) or "excel" in content_type or "spreadsheetml" in content_type

    if is_excel:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = []
        for idx, cell in enumerate(rows[0]):
            header = str(cell).strip() if cell not in (None, "") else f"column_{idx + 1}"
            headers.append(header)
        data_rows = []
        for row_values in rows[1:]:
            data_rows.append({headers[i]: (row_values[i] if row_values and i < len(row_values) else "") for i in range(len(headers))})
        return data_rows

    try:
        text = content.decode("utf-8-sig")
    except Exception:
        text = content.decode("latin-1")
    reader = csv.DictReader(io.StringIO(text))
    return list(reader)


def serialize_rows_to_csv(rows: List[dict], mockup_column: str = "mockup_url") -> str:
    if not rows:
        return ""
    headers = list(rows[0].keys())
    if mockup_column not in headers:
        headers.append(mockup_column)
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=headers)
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    return output.getvalue()


def process_job(
    job_id: str,
    rows: List[dict],
    template: dict,
    mapping: dict,
    skip_processed: bool = False,
    identifier_column: Optional[str] = None,
):
    session = SessionLocal()
    job = session.get(Job, job_id)
    if not job:
        session.close()
        return
    logger.info("Job %s started with %s rows", job_id, len(rows))
    total = len(rows)
    results = []
    mockup_column = "mockup_url"
    identifier_column = identifier_column or job.identifier_column
    processed_cache: Set[str] = set()
    if skip_processed and identifier_column:
        existing = session.query(ProcessedCompany.identifier).all()
        processed_cache = {row[0] for row in existing if row[0]}

    for idx, row in enumerate(rows):
        status = {"row": idx, "status": "pending", "url": None, "error": None}
        identifier_value = None
        if identifier_column:
            identifier_value = str(row.get(identifier_column) or "").strip()

        if skip_processed and identifier_value and identifier_value in processed_cache:
            status.update({"status": "skipped", "error": "Identifier already processed", "url": None})
            row[mockup_column] = row.get(mockup_column, "")
        else:
            try:
                image = apply_variables(template, row, mapping)
                url = upload_to_cloudinary(image, template["id"], idx, row)
                row[mockup_column] = url
                status["status"] = "done"
                status["url"] = url
                if identifier_value and identifier_value not in processed_cache:
                    processed_cache.add(identifier_value)
                    if not session.query(ProcessedCompany).filter_by(identifier=identifier_value).first():
                        session.add(ProcessedCompany(identifier=identifier_value))
            except Exception as e:
                status["status"] = "error"
                status["error"] = str(e)
                row[mockup_column] = ""

        results.append(status)
        job.progress = int(((idx + 1) / total) * 100)
        job.results = results
        stored_rows = job.rows or []
        if len(stored_rows) <= idx:
            stored_rows.append(row)
        else:
            stored_rows[idx] = row
        job.rows = stored_rows
        session.add(job)
        session.commit()
        logger.info("Job %s progress %s%% (%s/%s)", job_id, job.progress, idx + 1, total)

    # Ensure DB has the latest rows including mockup URLs
    job.rows = stored_rows
    job.status = "done"
    job.csv_path = None
    session.add(job)
    session.commit()
    logger.info("Job %s completed", job_id)
    session.close()


# === Routes ===
@app.on_event("startup")
async def seed_templates():
    ensure_upload_dir()
    Base.metadata.create_all(engine)
    run_migrations(engine)
    session = SessionLocal()
    try:
        has_templates = session.query(Template).first()
        if not has_templates:
            templates = [default_template()]
            roofing_path = os.path.join(BASE_DIR, "template", "roofing-template.jpg")
            if os.path.exists(roofing_path):
                templates.append(roofing_template())
            plumber_path = os.path.join(BASE_DIR, "template", "plumber-template-1.png")
            if os.path.exists(plumber_path):
                templates.append(plumber_template())
            for tpl in templates:
                session.add(
                    Template(
                        id=tpl["id"],
                        name=tpl["name"],
                        base_image_path=tpl["baseImagePath"],
                        overlays=tpl.get("overlays", []),
                        variables=tpl.get("variables", []),
                    )
                )
            session.commit()
    finally:
        session.close()


@app.get("/api/auth/has-users")
async def has_users(db: Session = Depends(get_session)):
    has_users = db.query(User).first() is not None
    return {"hasUsers": bool(has_users)}


@app.post("/api/auth/signup")
async def signup(payload: dict, db: Session = Depends(get_session)):
    email = (payload.get("email") or "").strip().lower()
    password = payload.get("password") or ""
    name = (payload.get("name") or "").strip()
    if not email or not password:
        raise HTTPException(status_code=400, detail="Email and password are required.")
    if get_user_by_email(db, email):
        raise HTTPException(status_code=400, detail="An account with that email already exists.")
    hashed = hash_password(password)
    user = User(email=email, name=name or email.split("@")[0], hashed_password=hashed)
    db.add(user)
    db.commit()
    db.refresh(user)
    token = create_access_token(user.id)
    return {"access_token": token, "token_type": "bearer", "user": user_to_dict(user)}


@app.post("/api/auth/login")
async def login(payload: dict, db: Session = Depends(get_session)):
    email = (payload.get("email") or "").strip().lower()
    password = payload.get("password") or ""
    user = get_user_by_email(db, email)
    if not user or not verify_password(password, user.hashed_password):
        raise HTTPException(status_code=401, detail="Invalid email or password.")
    token = create_access_token(user.id)
    return {"access_token": token, "token_type": "bearer", "user": user_to_dict(user)}


@app.get("/api/auth/me")
async def read_current_user(current_user: User = Depends(require_user)):
    return {"user": user_to_dict(current_user)}


@app.get("/api/templates")
async def list_templates(
    current_user: User = Depends(require_user),
    db: Session = Depends(get_session),
):
    templates = (
        db.query(Template)
        .filter(or_(Template.owner_id == current_user.id, Template.owner_id.is_(None)))
        .all()
    )
    return {"templates": [template_to_dict(t) for t in templates]}


@app.get("/api/templates/{template_id}")
async def get_template(
    template_id: str,
    current_user: User = Depends(require_user),
    db: Session = Depends(get_session),
):
    tpl = db.get(Template, template_id)
    if not tpl:
        raise HTTPException(status_code=404, detail="Template not found")
    if tpl.owner_id and tpl.owner_id != current_user.id:
        raise HTTPException(status_code=403, detail="You do not have access to this template.")
    return template_to_dict(tpl)


@app.post("/api/templates")
async def create_or_update_template(
    payload: dict,
    current_user: User = Depends(require_user),
    db: Session = Depends(get_session),
):
    tpl_id = payload.get("id") or uuid.uuid4().hex
    payload["id"] = tpl_id
    tpl = db.get(Template, tpl_id)
    if tpl:
        if tpl.owner_id and tpl.owner_id != current_user.id:
            raise HTTPException(status_code=403, detail="You do not have access to this template.")
        tpl.name = payload.get("name", tpl.name)
        tpl.base_image_path = payload.get("baseImagePath", tpl.base_image_path)
        tpl.variables = payload.get("variables", tpl.variables)
        tpl.masks = payload.get("masks", tpl.masks or [])
    else:
        tpl = Template(
            id=tpl_id,
            name=payload.get("name", "Untitled"),
            base_image_path=payload.get("baseImagePath", ""),
            variables=payload.get("variables", []),
            masks=payload.get("masks", []),
            owner_id=current_user.id,
        )
        db.add(tpl)
    db.commit()
    db.refresh(tpl)
    return {"template": template_to_dict(tpl)}


@app.post("/api/templates/upload-image")
async def upload_template_image(
    image: UploadFile = File(...),
    current_user: User = Depends(require_user),
):
    if not image.filename:
        raise HTTPException(status_code=400, detail="Missing filename")
    ensure_upload_dir()
    _, ext = os.path.splitext(image.filename)
    ext = ext if ext else ".jpg"
    filename = f"{uuid.uuid4().hex}{ext}"
    dest_path = os.path.join(UPLOADS_TEMPLATE_DIR, filename)
    with open(dest_path, "wb") as f:
        shutil.copyfileobj(image.file, f)
    rel_path = f"/uploads/templates/{filename}"
    return {"path": rel_path, "filename": filename}


@app.post("/api/templates/upload-mask")
async def upload_template_mask(
    mask: UploadFile = File(...),
    current_user: User = Depends(require_user),
):
    if not mask.filename:
        raise HTTPException(status_code=400, detail="Missing filename")
    content_type = mask.content_type or "image/png"
    try:
        content = await mask.read()
    except Exception:
        content = mask.file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty mask upload")
    data_url = f"data:{content_type};base64,{base64.b64encode(content).decode('utf-8')}"
    return {"data": data_url, "filename": mask.filename}


@app.post("/api/csv/inspect")
async def inspect_csv(
    csv_file: UploadFile = File(...),
    sample_size: int = Form(5),
    current_user: User = Depends(require_user),
):
    rows = parse_csv_file(csv_file)
    headers = list(rows[0].keys()) if rows else []
    return {"headers": headers, "sample_rows": rows[:sample_size]}


@app.post("/api/preview")
async def preview_mockups(
    template_id: str = Form(...),
    mapping: str = Form(...),
    limit: int = Form(3),
    csv_file: UploadFile = File(...),
    current_user: User = Depends(require_user),
    db: Session = Depends(get_session),
):
    try:
        mapping_dict = json.loads(mapping)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid mapping JSON")
    tpl_obj = db.get(Template, template_id)
    if not tpl_obj:
        raise HTTPException(status_code=404, detail="Template not found")
    if tpl_obj.owner_id and tpl_obj.owner_id != current_user.id:
        raise HTTPException(status_code=403, detail="You do not have access to this template.")
    tpl = template_to_dict(tpl_obj)
    rows = parse_csv_file(csv_file)
    previews = []
    for idx, row in enumerate(rows[:limit]):
        img = apply_variables(tpl, row, mapping_dict)
        buffer = io.BytesIO()
        img.save(buffer, format="JPEG", quality=70, optimize=True)
        buffer.seek(0)
        previews.append(
            {
                "row": idx,
                "image_base64": base64.b64encode(buffer.read()).decode("utf-8"),
                "row_data": row,
            }
        )
    return {"previews": previews}


@app.post("/api/batch")
async def start_batch(
    template_id: str = Form(...),
    mapping: str = Form(...),
    csv_file: UploadFile = File(...),
    skip_processed: bool = Form(False),
    identifier_column: Optional[str] = Form(None),
    current_user: User = Depends(require_user),
    db: Session = Depends(get_session),
):
    try:
        mapping_dict = json.loads(mapping)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid mapping JSON")
    tpl_obj = db.get(Template, template_id)
    if not tpl_obj:
        raise HTTPException(status_code=404, detail="Template not found")
    if tpl_obj.owner_id and tpl_obj.owner_id != current_user.id:
        raise HTTPException(status_code=403, detail="You do not have access to this template.")
    tpl = template_to_dict(tpl_obj)
    rows = parse_csv_file(csv_file)
    if not rows:
        raise HTTPException(status_code=400, detail="CSV has no rows")

    job_id = uuid.uuid4().hex
    job = Job(
        id=job_id,
        status="running",
        progress=0,
        results=[],
        rows=rows,
        mapping=mapping_dict,
        template_id=template_id,
        csv_path=None,
        skip_processed=bool(skip_processed),
        identifier_column=identifier_column,
        owner_id=current_user.id,
    )
    db.add(job)
    db.commit()
    # Run blocking job generation in a thread to keep the event loop responsive
    loop = asyncio.get_event_loop()
    loop.create_task(
        asyncio.to_thread(
            process_job,
            job_id,
            rows,
            tpl,
            mapping_dict,
            bool(skip_processed),
            identifier_column,
        )
    )
    return {"job_id": job_id, "total_rows": len(rows)}


@app.get("/api/jobs/{job_id}")
async def job_status(
    job_id: str,
    current_user: User = Depends(require_user),
    db: Session = Depends(get_session),
):
    job = db.get(Job, job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    if job.owner_id and job.owner_id != current_user.id:
        raise HTTPException(status_code=403, detail="You do not have access to this job.")
    # Derive progress from processed rows to avoid stale 0% readings
    derived_progress = job.progress or 0
    try:
        total = len(job.rows or [])
        processed = len(job.results or [])
        done_like = len([r for r in job.results or [] if r.get("status") in ("done", "skipped", "error")])
        derived_progress = int((max(processed, done_like) / total) * 100) if total else job.progress or 0
        if derived_progress > (job.progress or 0):
            job.progress = derived_progress
            db.add(job)
            db.commit()
            logger.info("Job %s derived progress updated to %s%% (processed=%s/%s)", job_id, derived_progress, processed, total)
    except Exception:
        pass
    data = job_to_dict(job)
    data["progress"] = max(data.get("progress") or 0, derived_progress or 0)
    logger.info("Job %s status polled: %s%%, stored=%s", job_id, data["progress"], job.progress)
    return data


@app.get("/api/jobs/{job_id}/csv")
async def download_job_csv(
    job_id: str,
    current_user: User = Depends(require_user),
    db: Session = Depends(get_session),
):
    job = db.get(Job, job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    if job.owner_id and job.owner_id != current_user.id:
        raise HTTPException(status_code=403, detail="You do not have access to this job.")
    if job.status != "done":
        raise HTTPException(status_code=400, detail="Job not finished")
    rows = job.rows or []
    # Fill in mockup URLs from results if rows are missing them
    for res in job.results or []:
        if res.get("status") == "done":
            row_idx = res.get("row")
            url = res.get("url")
            if row_idx is not None and url and row_idx < len(rows):
                rows[row_idx]["mockup_url"] = url
    csv_text = serialize_rows_to_csv(rows)
    filename = f"job_{job_id}.csv"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return Response(content=csv_text, media_type="text/csv", headers=headers)
